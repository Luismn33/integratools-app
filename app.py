from flask import Flask, render_template, request, redirect, url_for, send_from_directory
import qrcode
import subprocess
import os
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)

# Link del PDF para el QR (ajusta si usas QR)
pdf_url = ""
ruta_qr = os.path.join('static', 'qr', 'qr_code.png')

# Ruta base del proyecto
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def obtener_descripciones_excel():
    try:
        ruta_excel = os.path.join(BASE_DIR, "data", "INVENTARIO_MORATO_VALORIZADO.xlsx")
        wb = load_workbook(ruta_excel, data_only=True)
        ws = wb["CORRECTIVA"]
        
        fila_encabezados = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        print(f"Encabezados encontrados: {fila_encabezados}")  # Opcional para depuración

        try:
            idx_descripcion = fila_encabezados.index("DESCRIPCION")
        except ValueError:
            print(f"No se encontró la columna 'DESCRIPCION'.")
            return []

        descripciones = []
        for fila in ws.iter_rows(min_row=2, values_only=True):
            descripcion = fila[idx_descripcion]
            if descripcion:
                descripciones.append(str(descripcion).strip())

        return descripciones

    except Exception as e:
        print("Error al leer productos:", e)
        return []

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generar_qr')
def generar_qr():
    # Generar el link absoluto al PDF alojado localmente (servicio Flask)
    pdf_url = url_for('static', filename='pdf/correctiva_no1.pdf', _external=True) + f"?v={int(time.time())}"
    print("URL generada para el QR:", pdf_url)  # <-- Línea de depuración útil

    if os.path.exists(ruta_qr):
        os.remove(ruta_qr)
        
    # Crear el QR con la URL al PDF
    qr = qrcode.make(pdf_url)
    qr.save(ruta_qr)

    return redirect(url_for('ver_qr'))

@app.route('/ver_qr')
def ver_qr():
    import time
    return render_template('qr.html', timestamp=int(time.time()))

@app.route('/actualizar_pdf')
def actualizar_pdf():
    try:
        subprocess.run(["python", "generar_pdf.py"], check=True)
        return "PDF generado exitosamente"
    except Exception as e:
        return f"Error al generar PDF: {e}"

@app.route('/formulario')
def formulario():
    productos = obtener_descripciones_excel()
    return render_template("formulario_correctiva.html", descripciones=productos)

@app.route('/guardar_formulario', methods=['POST'])
def guardar_formulario():
    total = int(request.form.get("total"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Validación"

    encabezados = ["Descripción", "Cantidad", "Cant. Requerida", "Marca", "Referencia", "# de Activo", "Estado", "Fecha"]
    ws.append(encabezados)

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    for i in range(total):
        fila = [
            request.form.get(f"descripcion_{i}"),
            request.form.get(f"cantidad_{i}"),
            request.form.get(f"cant_requerida_{i}"),
            request.form.get(f"marca_{i}"),
            request.form.get(f"referencia_{i}"),
            request.form.get(f"activo_{i}"),
            request.form.get(f"estado_{i}"),
            fecha
        ]
        ws.append(fila)

    # Crear archivo con nombre único por fecha
    fecha_archivo = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo_guardado = os.path.join(BASE_DIR, "data", f"validacion_correctiva_{fecha_archivo}.xlsx")
    wb.save(archivo_guardado)

    return render_template("registro_exitoso.html", archivo=os.path.basename(archivo_guardado))

@app.route('/descargas/<archivo>')
def descargar_archivo(archivo):
    ruta_descargas = os.path.join(BASE_DIR, "data")
    return send_from_directory(ruta_descargas, archivo, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)